import pickle
import random
from collections import defaultdict
import string
import wiktionary
import openpyxl
from sklearn.model_selection import train_test_split

"""
def get_supersenses_from_sequoia_lemma_in_sentence(lemma, sentence):
    supersenses_to_be_returned = set()
    index2supersense = {}
    index_patterns = [f"{i}:" for i in range(1, 20)]
    indices = [f"{i}" for i in range(1, 20)]

    for word in sentence:
        if word[2] == lemma:
            if ';' in word[11]:
                supersenses = word[11].strip().split(";")
            else:
                supersenses = [word[11]]

            for supersense in supersenses:

                if any(pattern in supersense for pattern in index_patterns):
                    num = next(pattern for pattern in index_patterns if pattern in supersense)
                    supersenses_to_be_returned.add(supersense.strip(num))
                    index2supersense[num.strip(':')] = supersense.strip(num)

                elif any(index in supersense for index in indices):
                    num = next(index for index in indices if index in supersense)
                    if num in index2supersense:
                        supersenses_to_be_returned.add(index2supersense[num])
                        index2supersense[num.strip(':')] = supersense.strip(num)
                    else:
                        for word in sentence:
                            if ';' in word[11]:
                                ss = word[11].strip().split(";")
                            else:
                                ss = [word[11]]
                            for s in ss:
                                if any(pattern in s for pattern in index_patterns):
                                    num = next(pattern for pattern in index_patterns if pattern in s)
                                    supersenses_to_be_returned.add(s.strip(num))
                                    index2supersense[num.strip(':')] = s.strip(num)
                else:
                    supersenses_to_be_returned.add(supersense)
    # print(list(supersenses_to_be_returned))
    return list(supersenses_to_be_returned)


def get_noun_entities_from_corpus(old_seeds_file, sequoia_file, wiki_file):
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    seen_words = set()
    pages = wiki.pages
    lexical_entries = wiki.lexical_entries
    lexical_senses = wiki.lexical_senses
    ids_definitions_missing = set()
    ids_examples_missing = set()
    lemma2supersenses = defaultdict(set)

    with open(old_seeds_file, "r", encoding="utf-8") as z:
        z.readline()
        while 1:
            line = z.readline()
            if not line:
                break
            sline = line.strip().strip("\n").split("\t")
            seen_words.add(sline[0])

    corpus_name = "sequoia"
    new_seeds = open(f"seeds_from_{corpus_name}.txt", 'w', encoding="utf-8")
    corpus_name_file = f"seeds_from_{corpus_name}.txt"
    corpus = open(sequoia_file, 'r', encoding="utf-8")
    space_pattern = "\t"

    new_seeds.write("LEMMA" + space_pattern)
    new_seeds.write("DEFINITION" + space_pattern)
    for i in range(1, 11):
        new_seeds.write(f"EXAMPLE_{i}" + space_pattern)
    new_seeds.write("\n")
    lemma2definitions = defaultdict(list)
    lemma2examples = defaultdict(list)
    lemmas = set()
    last_line = False
    very_first_word = True
    words = []
    while 1:
        line = corpus.readline()
        if line.startswith("#"):
            continue
        if not line:
            last_line = True

        if line.strip():
            sline = line.strip().strip("\n").split("\t")
            if (sline[0] == "1" or last_line) and not very_first_word:
                for word in words:
                    if word[3] == "N":
                        lemma = word[2]
                        lemmas.add(lemma)
                        supersenses = get_supersenses_from_sequoia_lemma_in_sentence(lemma, words)
                        # print(supersenses)
                        for supersense in supersenses:
                            if supersense != "*":
                                lemma2supersenses[lemma].add(supersense)
                # get current sentence 1st word
                words = [sline]
            else:
                sline = line.strip().strip("\n").split("\t")
                words.append(sline)
                very_first_word = False
        if last_line:
            break

    for lemma in list(lemmas):
        page_id = lemma
        if page_id in pages and page_id not in seen_words:
            seen_words.add(page_id)
            entry_num = 0
            for entry_id in pages[page_id].get_entry_ids():
                if entry_id in lexical_entries:
                    sense_num = 0
                    entry_num += 1
                    le = lexical_entries[entry_id]
                    lemma = le.get_lemma()
                    for sense_id in lexical_entries[entry_id].get_sense_ids():
                        if sense_id in lexical_senses:
                            sense_num += 1
                            ls = lexical_senses[sense_id]
                            new_sense_id = f"{lemma}__nom_{entry_num}_sens_{sense_num}"
                            is_ambiguous = "yes" if (pages[page_id].is_multi_entries_page() or lexical_entries[
                                entry_id].is_multi_senses_entry()) else "no"
                            new_seeds.write(f"{lemma}{space_pattern}")

                            if ls.get_definition() is None:
                                ids_definitions_missing.add(sense_id)
                                # new_seeds.write(f"DEF:{temp_definitions[sense_id]}{space_pattern}")
                                new_seeds.write(f"DEF:{''}{space_pattern}")
                            else:
                                new_seeds.write(f"DEF:{ls.get_definition()}{space_pattern}")

                            for example in ls.get_examples():
                                if example == "":
                                    ids_examples_missing.add(sense_id)
                                else:
                                    new_seeds.write(f"EX:{example}{space_pattern}")
                            new_seeds.write("\n")

    new_seeds.close()
    corpus.close()
    print("DEF MANQUANTES: ", len(ids_definitions_missing))
    print("EX MANQUANTS: ", len(ids_examples_missing))
    print(f"len(lemmas) = {len(lemmas)}")
    return corpus_name_file, lemma2supersenses


def write_new_seeds(corpus_seeds_file, lemma2supersenses):
    new_seeds = open("new_seeds.txt", 'w', encoding="utf-8")
    corpus = open(corpus_seeds_file, 'r', encoding="utf-8")
    space_pattern = "\t"
    new_seeds.write("LEMMA" + space_pattern)
    new_seeds.write("SUPERSENSES" + space_pattern)
    new_seeds.write("DEFINITION" + space_pattern)
    for i in range(1, 11):
        new_seeds.write(f"EXAMPLE_{i}" + space_pattern)
    new_seeds.write("\n")
    nb_examples = 10
    senses = []
    lemma = ""
    first_line = True
    last_line = False
    while 1:
        line = corpus.readline()

        if line.startswith("LEMMA"):
            continue
        if not line:
            last_line = True

        sense = line.strip().split('\t')
        sense = sense[:-1] if sense[-1] == "\n" else sense
        if (sense[0] == lemma or first_line) and not last_line:
            senses.append(sense)
            if first_line:
                lemma = sense[0]
                first_line = False
        if sense[0] != lemma or last_line:
            # process every line for previous lemma
            lemma_first_time = True
            supersenses_not_written = True
            for line in senses:
                lemma = line[0]
                supersenses = list(lemma2supersenses[lemma])
                examples = []
                definition = line[1]
                for ex in line[2:]:
                    examples.append(ex)

                new_seeds.write(f"{lemma}{space_pattern}") if lemma_first_time else new_seeds.write(f"{space_pattern}")

                if supersenses_not_written:
                    supersenses_list = ""
                    for ss in supersenses:
                        supersenses_list += f"{ss};"
                    supersenses_list = supersenses_list.strip(';')
                    new_seeds.write(f"{supersenses_list}{space_pattern}")
                else:
                    new_seeds.write(f"{space_pattern}")
                supersenses_not_written = False
                new_seeds.write(f"{definition}{space_pattern}")
                for ex in examples:
                    new_seeds.write(f"{ex}{space_pattern}")
                new_seeds.write("\n")
                lemma_first_time = False

            senses = [sense]
            lemma = sense[0]

        if last_line:
            break

    new_seeds.close()
    corpus.close()


def from_table_to_text_file(table_file, supersenses_to_consider):
    table_name = ""
    for c in table_file:
        if c == ".":
            break
        else:
            table_name += c

    with open(f'{table_name}.txt', 'w', encoding="utf-8") as file:
        wb = openpyxl.load_workbook(table_file)
        ws = wb.active
        headers = []

        first_line = ""
        for cell in ws.iter_cols(min_row=1, max_row=1, values_only=True):
            if cell[0] is not None:
                headers.append(str(cell[0]).strip().lower())
                first_line += f"{str(cell[0]).strip().lower()}\t"
        first_line = first_line.strip("\t")
        file.write(first_line + "\n")

        for row in ws.iter_rows(min_row=2):
            line = ""
            row_supersense = None
            for j, cell in enumerate(row):
                if j < len(headers):
                    if cell.value is None:
                        line += f"\t"
                    else:
                        value = str(cell.value).strip()
                        if headers[j].lower() == 'supersense':
                            if value in supersenses_to_consider:
                                row_supersense = value
                                line += f"{value}\t"
                            else:
                                break
                        else:
                            line += f"{value}\t"
            if row_supersense:
                line = line.strip("\t")
                file.write(f"{line}\n")
        print(headers)


def get_sense_id_and_entry_id_from_lemma_and_defition(lemma, definition, wiki):
    for entry_id in wiki.pages[lemma].entry_ids:
        for sense_id in wiki.lexical_entries[entry_id].sense_ids:
            sense = wiki.lexical_senses[sense_id]
            if sense.definition == definition.replace("DEF:", ""):
                le_id = entry_id
                ls_id = sense_id
    return ls_id, le_id


def add_sense_id_and_entry_id_to_seeds_checked(file="seeds_finalV.txt", wiki_file="wiki_dump.pkl"):
    old_file = open(file, 'r', encoding="utf-8")
    new_file = open("seeds.txt", 'w', encoding="utf-8")

    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)

    new_file.write(old_file.readline().replace("\tquestion", "").replace("id_sense\t", ""))
    while 1:
        line = old_file.readline()
        if not line:
            break
        sline = line.strip().strip("\n").split("\t")
        print(sline)
        lemma = sline[0]
        if len(sline) <= 7:
            definition = "DEF:"
        else:
            definition = sline[7]
        if not sline[2] == "oui":
            if sline[1] == "" and sline[5] == "" and sline[6] == "":
                sense_id, entry_id = get_sense_id_and_entry_id_from_lemma_and_defition(lemma, definition, wiki)
                new_sline = [el for el in sline]
                new_sline[5] = entry_id
                new_sline[6] = sense_id
                new_sline.pop(1)
                new_line = ""
                for el in new_sline:
                    new_line += f"{el}\t"
                new_line = new_line.strip("\t")
                new_line += "\n"
                new_line = new_line.replace("\t\t\t", "\t\t")
                new_file.write(new_line)
            else:
                new_sline = sline
                new_sline.pop(1)
                new_line = ""
                for el in new_sline:
                    new_line += f"{el}\t"
                new_line = new_line.strip("\t")
                new_line += "\n"
                new_line = new_line.replace("\t\t", "\t")
                new_file.write(new_line)

    new_file.close()
    old_file.close()


def get_natural_selection_of_senses_from_corpus(wiki_file="wiki_dump.pkl", seeds_file="seeds_checked_V3.txt"):
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    nb_senses_hab = 0
    nb_senses = 0
    for sense_id in wiki.lexical_senses:
        if wiki.lexical_senses[sense_id].definition:
            if wiki.lexical_senses[sense_id].definition.startswith("Habitant de") or wiki.lexical_senses[
                sense_id].definition.startswith("Habitant d'") or wiki.lexical_senses[sense_id].definition.startswith(
                    "Habitant du") or wiki.lexical_senses[sense_id].definition.startswith("Habitante du") or \
                    wiki.lexical_senses[sense_id].definition.startswith("Habitante de") or wiki.lexical_senses[
                sense_id].definition.startswith("Habitante d'"):
                nb_senses_hab += 1
                nb_senses += 1
            else:
                nb_senses += 1
    seeds = open(seeds_file, 'r', encoding="utf-8")
    senses = open("baseline_random_senses_to_annotate_from_wiktionary.txt", 'w', encoding="utf-8")
    seeds_senses = {}

    first_line = seeds.readline()
    headers = first_line.strip().strip("\n").split("\t")
    senses.write(first_line.replace("\tambiguous", "").replace("\tdefinition\t", "\tlabels\tsynonyms\tdefinition\t"))

    while 1:
        line = seeds.readline()
        if not line:
            break
        if line.strip():
            sline = line.strip().strip("\n").split("\t")
            for i, header in enumerate(headers):
                if header == "supersense":
                    supersense = sline[i]
                if header == "id_sense_wiki":
                    sense_id = sline[i]
            seeds_senses[sense_id] = supersense

    lexical_senses_ids = list(wiki.lexical_senses.keys())
    # random.shuffle(lexical_senses_ids)
    # random_senses_ids = lexical_senses_ids[:1000]

    random_senses_ids = random.sample(lexical_senses_ids, 1000)

    for lemma in wiki.pages:
        for entry_id in wiki.pages[lemma].entry_ids:
            if wiki.lexical_entries[entry_id].sense_ids:
                for sense_id in wiki.lexical_entries[entry_id].sense_ids:
                    if sense_id in random_senses_ids:

                        labels = ""
                        if wiki.lexical_senses[sense_id].labels:
                            for label in list(wiki.lexical_senses[sense_id].labels):
                                labels += (label + ";")
                        labels = labels.strip(";")

                        synonyms = ""
                        if wiki.lexical_senses[sense_id].synonyms:
                            for synonym in list(wiki.lexical_senses[sense_id].synonyms):
                                synonyms += (synonym + ";")
                        synonyms = synonyms.strip(";")

                        definition = wiki.lexical_senses[sense_id].definition

                        examples = ""
                        for example in wiki.lexical_senses[sense_id].examples:
                            examples += (example + "\t")
                        examples = examples.strip("\t")

                        if sense_id in seeds_senses:
                            senses.write(
                                f"{lemma}\t{seeds_senses[sense_id]}\t{entry_id}\t{sense_id}\t{labels}\t{synonyms}\t{definition}\t{examples}\n")
                        else:
                            senses.write(
                                f"{lemma}\t{'*'}\t{entry_id}\t{sense_id}\t{labels}\t{synonyms}\t{definition}\t{examples}\n")

    seeds.close()
    senses.close()
    print(nb_senses_hab)
    print(nb_senses)

# supersenses acknowleged
SUPERSENSES = ['act', 'animal', 'artifact', 'attribute', 'body', 'cognition',
               'communication', 'event', 'feeling', 'food', 'institution', 'act*cognition',
               'object', 'possession', 'person', 'phenomenon', 'plant', 'artifact*cognition',
               'quantity', 'relation', 'state', 'substance', 'time', 'groupxperson']


def get_new_seeds(old_seeds, dev_test_sets_file, wiki_dump):
    with open(wiki_dump, "rb") as file:
      wiki = pickle.load(file)

    with open(dev_test_sets_file, 'r', encoding="utf-8") as file:
        dev_test_senses_ids = []
        headers = file.readline()
        old_seeds_lines = file.readlines()
        for line in old_seeds_lines:
            for i, el in enumerate(line.strip("\n").strip().split("\t")):
                if headers[i] == "id_sense_wiki":
                    dev_test_senses_ids.append(el)

    new_seeds = open("new_seeds_V2.txt", 'w', encoding="utf-8")
    with open(old_seeds, 'r', encoding="utf-8") as file:
        new_lines = []
        old_seeds_lines = file.readlines()
        headers = old_seeds_lines[0].strip().strip("\n").split("\t")
        for line in old_seeds_lines:
            for i, el in enumerate(line.strip("\n").strip().split("\t")):
                if headers[i] == "id_sense_wiki":
                    if el not in dev_test_senses_ids:
                        new_lines.append(line)

        nb = 0
        gentiles = []

        for page_id in wiki.pages:
            for entry_id in wiki.pages[page_id].entry_ids:
                if wiki.lexical_entries[entry_id].sense_ids:
                    for sense_id in wiki.lexical_entries[entry_id].sense_ids:
                        if wiki.lexical_senses[sense_id].definition:
                            if sense_id not in dev_test_senses_ids and wiki.lexical_senses[sense_id].definition.startswith("Habitant"):
                                gentiles.append((sense_id, entry_id, page_id))
        gentiles_to_add_to_seeds = random.sample(gentiles, 50)
        for line in new_lines:
            new_seeds.write(line)
        for sense_id_entry_id_lemma in gentiles_to_add_to_seeds:
            sense_id = sense_id_entry_id_lemma[0]
            entry_id = sense_id_entry_id_lemma[1]
            lemma = sense_id_entry_id_lemma[2]
            supersense = "person"
            labels = wiki.lexical_senses[sense_id].labels
            examples =  wiki.lexical_senses[sense_id].examples
            definition = wiki.lexical_senses[sense_id].definition
            line_to_write = ""
            line_to_write += (lemma + "\t")
            line_to_write += ("" + "\t")
            line_to_write += (supersense + "\t")
            line_to_write += (entry_id + "\t")
            line_to_write += (sense_id + "\t")
            line_to_write += (definition + "\t")
            for example in examples:
                line_to_write += (example + "\t")
            line_to_write.strip("\t")
            line_to_write += "\n"

            new_seeds.write(line_to_write)

    new_seeds.close()

def keep_n_elements(lst, n):
    random.shuffle(lst)
    if len(lst) >= n:
        return lst[:n]
    else:
        return lst


def split_data(data_list):
    train_data, test_data = train_test_split(data_list, test_size=0.2)
    dev_data, test_data = train_test_split(test_data, test_size=0.5)
    return train_data, dev_data, test_data


def get_data_sets(n=1, seeds_file_path="seeds.txt", wiki_file="wiki_dump.pkl"):
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    for i in range(1, n+1):
        with open(seeds_file_path, 'r', encoding="utf-8") as seeds_file:
            id2data = defaultdict(dict)
            supersense2ids = defaultdict(list)
            train = []
            dev = []
            test = []
            # build dictionnary structures
            headers = seeds_file.readline().strip("\n").strip().split("\t")
            id = 0
            while 1:
                id += 1
                line = seeds_file.readline()
                if not line:
                    break
                sline = line.strip("\n").strip().split("\t")
                examples = []
                labels = []
                for header, value in zip(headers, sline):
                    if header == "id_sense_wiki":
                        sense_id = value
                        if wiki.lexical_senses[sense_id].labels:
                            labels = list(wiki.lexical_senses[sense_id].labels)
                    if header == "supersense":
                        supersense = value
                    if header == "lemma":
                        lemma = value
                    if header == "definition":
                        value = value.replace("DEF:", "")
                        definition = value
                        definition_with_lemma = f"{lemma} : {value}"
                        definition_with_labels = ""
                        if labels:
                            for label in labels:
                                definition_with_labels += f"({label}) "
                            definition_with_labels += value
                        else:
                            definition_with_labels = value
                        definition_with_lemma_and_labels = f"{lemma} : "
                        if labels:
                            for label in labels:
                                definition_with_lemma_and_labels += f"({label}) "
                            definition_with_lemma_and_labels += value
                        else:
                            definition_with_lemma_and_labels = f"{lemma} : {value}"
                    if "example" in header:
                        if value:
                            value = value.replace("EX:", "")
                            examples.append(value)
                if supersense in SUPERSENSES:
                    id2data[id]["definition"] = (definition, supersense)
                    id2data[id]["definition_with_lemma"] = (definition_with_lemma, supersense)
                    id2data[id]["definition_with_labels"] = (definition_with_labels, supersense)
                    id2data[id]["definition_with_lemma_and_labels"] = (definition_with_lemma_and_labels, supersense)
                    id2data[id]["examples"] = (examples, supersense)
                    supersense2ids[supersense].append(id)

            # build id lists for classification sets
            for supersense in supersense2ids:
                train_ids, dev_ids, test_ids = split_data(supersense2ids[supersense])
                train += train_ids
                dev += dev_ids
                test += test_ids

            # serialize the different structures created
            with open(f"{i}_id2data.pkl", "wb") as file:
                pickle.dump(id2data, file)
            with open(f"{i}_supersense2ids.pkl", "wb") as file:
                pickle.dump(supersense2ids, file)
            with open(f"{i}_train.pkl", "wb") as file:
                pickle.dump(train, file)
            with open(f"{i}_dev.pkl", "wb") as file:
                pickle.dump(dev, file)
            with open(f"{i}_test.pkl", "wb") as file:
                pickle.dump(test, file)


def get_fixed_nb_examples_data_sets(n=100, seeds_file="seeds_checked_V3.txt"):
    id2def_supersense = {}
    id2defwithlemma_supersense = {}
    supersense2ids = defaultdict(list)
    train = []
    dev = []
    test = []
    seeds_file = open(seeds_file, 'r', encoding="utf-8")
    # build dictionnary structures
    headers = seeds_file.readline().strip("\n").strip().split("\t")
    id = 0
    while 1:
        id += 1
        line = seeds_file.readline()
        if not line:
            break
        sline = line.strip("\n").strip().split("\t")
        for header, value in zip(headers, sline):
            if header == "supersense":
                supersense = value
            if header == "lemma":
                lemma = value
            if header == "definition":
                value = value.replace("DEF:", "")
                definition = value
                definitionwithlemma = f"{lemma} : {value}"
        if supersense in SUPERSENSES:
            id2def_supersense[id] = (definition, supersense)
            id2defwithlemma_supersense[id] = (definitionwithlemma, supersense)
            supersense2ids[supersense].append(id)

    # build id lists for classification sets
    for supersense in supersense2ids:

        train_ids, dev_ids, test_ids = split_data(keep_n_elements(supersense2ids[supersense], n=n))
        train += train_ids
        dev += dev_ids
        test += test_ids

    # serialize the different structures created
    with open(f"{n}_id2def_supersense.pkl", "wb") as file:
        pickle.dump(id2def_supersense, file)
    with open(f"{n}_id2defwithlemma_supersense.pkl", "wb") as file:
        pickle.dump(id2defwithlemma_supersense, file)
    with open(f"{n}_supersense2ids.pkl", "wb") as file:
        pickle.dump(supersense2ids, file)
    with open(f"{n}_train.pkl", "wb") as file:
        pickle.dump(train, file)
    with open(f"{n}_dev.pkl", "wb") as file:
        pickle.dump(dev, file)
    with open(f"{n}_test.pkl", "wb") as file:
        pickle.dump(test, file)
    seeds_file.close()
        

def new_eval_data(nouns_file='/home/nangleraud/PycharmProjects/stageM1/common_lc_singular.txt', wiki_file='/home/nangleraud/PycharmProjects/stageM1/wiki_dump.pkl'):

    alphabet = list(string.ascii_lowercase)

    new_data = open("/home/nangleraud/PycharmProjects/stageM1/eval_data.txt", 'w', encoding="utf-8")
    nouns = []

    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
        pages = wiki.pages

    with open(nouns_file, 'r', encoding="utf-8") as file:
        lines = file.readlines()

    for line in lines:
        sline = line.strip()
        if sline in (alphabet+['%', '*', '$']):
            continue
        else:
            if (not sline.endswith('s')) and (not sline.endswith('x')):
                nouns.append(sline)
            else:
                if sline[:-1] in pages:
                    nouns.append(sline[:-1])
                else:
                    nouns.append(sline)

    random.shuffle(nouns)

    for noun in nouns:
        lemma = noun
        if noun in pages:
            if pages[noun].entry_ids:
                entry_id = random.choice(pages[noun].entry_ids)
                entry = wiki.lexical_entries[entry_id]
                if entry.sense_ids:
                    sense_id = random.choice(entry.sense_ids)
                    sense = wiki.lexical_senses[sense_id]
                    definition = sense.definition
                    if definition:

                        examples = sense.examples
                        sexamples = ""
                        for example in examples:
                            sexamples += example
                            sexamples += "\t"
                        sexamples = sexamples.strip("\t")

                        labels = sense.labels
                        slabels = ""
                        if labels:
                            for label in labels:
                                slabels += label
                                slabels += ";"
                            slabels = slabels.strip(";")

                        synonyms = sense.synonyms
                        ssynonyms = ""
                        if synonyms:
                            for synonym in synonyms:
                                ssynonyms += synonym
                                ssynonyms += ";"
                            ssynonyms = ssynonyms.strip(";")

                        new_data.write(lemma + "\t")
                        new_data.write("\t")
                        new_data.write(entry_id + "\t")
                        new_data.write(sense_id + "\t")
                        new_data.write(slabels + "\t")
                        new_data.write(ssynonyms + "\t")
                        new_data.write(definition + "\t")
                        new_data.write(sexamples)
                        new_data.write("\n")

    new_data.close()



def eval_data_1200(eval_file='eval_data.txt'):
    with open(eval_file, 'r', encoding="utf-8") as file:
        lines = file.readlines()

    file = open("eval_data_1200.txt", 'w', encoding="utf-8")

    for line in lines[:1201]:
        file.write(line)

    file.close()





def new_train():
    train_file = open("new_seeds_V2.txt", 'r', encoding="utf-8")
    new_train_file = open("train_data.txt", 'w', encoding="utf-8")

    lines = train_file.readlines()
    train_headers = lines[0]
    print(train_headers)
    train_headers = train_headers.strip("\t").split('\t')
    print(train_headers)

    for i, header in enumerate(train_headers):
        if header == "ambiguous":
            supp_index = i
    first_line = ""
    for header in train_headers:
        if header != "ambiguous":
            first_line += (header + "\t")
    new_train_file.write(first_line)

    for line in lines[1:]:
        els = line.strip("\t").split("\t")
        els.pop(supp_index)
        print(els)
        new_line = ""
        for el in els:
            el = el.strip().strip("DEF:").strip("EX:")
            new_line += el
            new_line += "\t"
        new_train_file.write(new_line)

    train_file.close()
    new_train_file.close()


def eval_data_1200(eval_file='eval_data.txt'):
    with open(eval_file, 'r', encoding="utf-8") as file:
        lines = file.readlines()

    file = open("eval_data_1200.txt", 'w', encoding="utf-8")

    for line in lines[:1201]:
        file.write(line)

    file.close()


eval_data_1200()


import pandas as pd

# Define the input XLSX file and output TXT file paths
xlsx_file = 'eval_data_1200_annotated.xlsx'
txt_file = 'eval_data.txt'

# Read the XLSX file into a pandas DataFrame
df = pd.read_excel(xlsx_file, engine='openpyxl')

# Filter rows where the "commentaire" column is equal to "ok"
df_filtered = df[df['commentaire'] == 'ok']

# Create a tab-separated TXT file with the same structure
df_filtered.to_csv(txt_file, sep='\t', index=False, header=True)

print(f"Filtered data has been saved to {txt_file}")


old_eval_data = open("eval_data.txt", 'r', encoding="utf-8")
train_data = open("new_seeds_V2.txt", 'r', encoding="utf-8")
eval_data = open("new_train.txt", 'r', encoding="utf-8")

old_eval_sense_ids = []
eval_sense_ids = []
train_sense_ids = []
supp_sense_ids = []

eval_lines = eval_data.readlines()
headers = eval_lines[0].strip().split("\t")
for line in eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                eval_sense_ids.append(el)

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                train_sense_ids.append(el)

old_eval_lines = old_eval_data.readlines()
headers = old_eval_lines[0].strip().split("\t")
for line in old_eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                old_eval_sense_ids.append(el)

for sense_id in old_eval_sense_ids:
    if sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)
    elif sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

eval_data.close()
train_data.close()
old_eval_data.close()

print(len(supp_sense_ids))
"""
"""
old_eval_data = open("eval_data.txt", 'r', encoding="utf-8")
train_data = open("new_seeds_V2.txt", 'r', encoding="utf-8")
eval_data = open("new_train.txt", 'r', encoding="utf-8")

old_eval_sense_ids = []
eval_sense_ids = []
train_sense_ids = []
supp_sense_ids = []

eval_lines = eval_data.readlines()
headers = eval_lines[0].strip().split("\t")
for line in eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                eval_sense_ids.append(el)

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                train_sense_ids.append(el)

old_eval_lines = old_eval_data.readlines()
headers = old_eval_lines[0].strip().split("\t")
for line in old_eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                old_eval_sense_ids.append(el)

for sense_id in old_eval_sense_ids:
    if sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)
    elif sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

for sense_id in eval_sense_ids:
    if sense_id in old_eval_sense_ids:
        supp_sense_ids.append(sense_id)
    elif sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

for sense_id in train_sense_ids:
    if sense_id in old_eval_sense_ids:
        supp_sense_ids.append(sense_id)
    elif sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)

eval_data.close()
train_data.close()
old_eval_data.close()

print(len(supp_sense_ids))
"""
"""
with open("wiki_dump.pkl", "rb") as file:
    wiki = pickle.load(file)

new_train_data = open("new_train.txt", 'r', encoding="utf-8")
train_data = open("new_seeds_V2.txt", 'r', encoding="utf-8")

sense_ids = []
entry_ids = []
lemmas = []
supersenses = []

new_train_lines = new_train_data.readlines()
headers = new_train_lines[0].strip().split("\t")
for line in new_train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                sense_ids.append(el)
            if headers[i] == "lemma":
                lemmas.append(el)
            if headers[i] == "id_entry_wiki":
                entry_ids.append(el)
            if headers[i] == "supersense":
                supersenses.append(el)

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                sense_ids.append(el)
            if headers[i] == "lemma":
                lemmas.append(el)
            if headers[i] == "id_entry_wiki":
                entry_ids.append(el)
            if headers[i] == "supersense":
                supersenses.append(el)

train_final = open("train_data.txt", 'w', encoding="utf-8")

for lemma, entry_id, sense_id, supersense in zip(lemmas, entry_ids, sense_ids, supersenses):
    sense = wiki.lexical_senses[sense_id]
    definition = sense.definition
    if definition:
        examples = sense.examples
        sexamples = ""
        for example in examples:
            sexamples += example
            sexamples += "\t"
        sexamples = sexamples.strip("\t")

        labels = sense.labels
        slabels = ""
        if labels:
            for label in labels:
                slabels += label
                slabels += ";"
            slabels = slabels.strip(";")

        synonyms = sense.synonyms
        ssynonyms = ""
        if synonyms:
            for synonym in synonyms:
                ssynonyms += synonym
                ssynonyms += ";"
            ssynonyms = ssynonyms.strip(";")

        train_final.write(lemma + "\t")
        train_final.write("ok\t")
        train_final.write(supersense + "\t")
        train_final.write(entry_id + "\t")
        train_final.write(sense_id + "\t")
        train_final.write(slabels + "\t")
        train_final.write(ssynonyms + "\t")
        train_final.write(definition + "\t")
        train_final.write(sexamples)
        train_final.write("\n")

train_final.close()
"""

"""
with open("/home/nangleraud/PycharmProjects/stageM1/wiki_dump.pkl", "rb") as file:
    wiki = pickle.load(file)

new_train_data = open("/home/nangleraud/PycharmProjects/stageM1/new_train.txt", 'r', encoding="utf-8")
train_data = open("/home/nangleraud/PycharmProjects/stageM1/new_seeds_V2.txt", 'r', encoding="utf-8")

sense_ids = []
entry_ids = []
lemmas = []
supersenses = []

new_train_lines = new_train_data.readlines()
headers = new_train_lines[0].strip().split("\t")
for line in new_train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                sense_ids.append(el)
            if headers[i] == "lemma":
                lemmas.append(el)
            if headers[i] == "id_entry_wiki":
                entry_ids.append(el)
            if headers[i] == "supersense":
                supersenses.append(el)

print(len(sense_ids))

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                sense_ids.append(el)
            if headers[i] == "lemma":
                lemmas.append(el)
            if headers[i] == "id_entry_wiki":
                entry_ids.append(el)
            if headers[i] == "supersense":
                supersenses.append(el)
print(len(sense_ids))
train_final = open("/home/nangleraud/PycharmProjects/stageM1/train_data.txt", 'w', encoding="utf-8")

for lemma, entry_id, sense_id, supersense in zip(lemmas, entry_ids, sense_ids, supersenses):
    sense = wiki.lexical_senses[sense_id]
    definition = sense.definition
    if definition:
        examples = sense.examples
        sexamples = ""
        for example in examples:
            sexamples += example
            sexamples += "\t"
        sexamples = sexamples.strip("\t")

        labels = sense.labels
        slabels = ""
        if labels:
            for label in labels:
                slabels += label
                slabels += ";"
            slabels = slabels.strip(";")

        synonyms = sense.synonyms
        ssynonyms = ""
        if synonyms:
            for synonym in synonyms:
                ssynonyms += synonym
                ssynonyms += ";"
            ssynonyms = ssynonyms.strip(";")

        train_final.write(lemma + "\t")
        train_final.write("ok\t")
        train_final.write(supersense + "\t")
        train_final.write(entry_id + "\t")
        train_final.write(sense_id + "\t")
        train_final.write(slabels + "\t")
        train_final.write(ssynonyms + "\t")
        train_final.write(definition + "\t")
        train_final.write(sexamples)
        train_final.write("\n")

train_final.close()
"""

"""
with open("train.pkl", 'rb') as file:
    train_examples = pickle.load(file)
with open("dev.pkl", 'rb') as file:
    dev_examples = pickle.load(file)
with open("test.pkl", 'rb') as file:
    test_examples = pickle.load(file)


    for dev in dev_examples:
        for test in test_examples:
            if tr['definition'] == dev['definition']:
                print(tr['definition'])
            if tr['definition'] == test['definition']:
                print(tr['definition'])
            if test['definition'] == dev['definition']:
                print(test['definition'])
"""
"""
train_data = open("train_data.txt", 'r', encoding="utf-8")
eval_data = open("eval_data.txt", 'r', encoding="utf-8")

eval_sense_ids = []
train_sense_ids = []
supp_sense_ids = []

eval_lines = eval_data.readlines()
headers = eval_lines[0].strip().split("\t")
for line in eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                if el in eval_sense_ids:
                    print(el)
                eval_sense_ids.append(el)

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                train_sense_ids.append(el)

for sense_id in eval_sense_ids:
    if sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

for sense_id in train_sense_ids:
    if sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)

eval_data.close()
train_data.close()

print(len(supp_sense_ids))

print(len(train_sense_ids))
print(len(set(train_sense_ids)))

print(len(eval_sense_ids))
print(len(set(eval_sense_ids)))
"""

"""

with open("train.pkl", 'rb') as file:
    train_examples = pickle.load(file)
with open("dev.pkl", 'rb') as file:
    dev_examples = pickle.load(file)
with open("test.pkl", 'rb') as file:
    test_examples = pickle.load(file)

print(len(train_examples))
print(len(dev_examples))
print(len(test_examples))


for tr in train_examples:
    for dev in dev_examples:
        if tr['definition'] == dev['definition']:
            print(tr['definition'])
        for test in test_examples:
            if tr['definition'] == test['definition']:
                print(tr['definition'])
            if test['definition'] == dev['definition']:
                print(test['definition'])
"""
"""
train_data = open("train_data.txt", 'r', encoding="utf-8")
eval_data = open("eval_data.txt", 'r', encoding="utf-8")
test_data = open("test_set_2.txt", 'r', encoding="utf-8")

eval_sense_ids = []
train_sense_ids = []
test_sense_ids = []
supp_sense_ids = []

eval_lines = eval_data.readlines()
headers = eval_lines[0].strip().split("\t")
for line in eval_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                eval_sense_ids.append(el)

train_lines = train_data.readlines()
headers = train_lines[0].strip().split("\t")
for line in train_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                train_sense_ids.append(el)

test_lines = test_data.readlines()
headers = test_lines[0].strip().split("\t")
for line in test_lines[1:]:
    for i, el in enumerate(line.strip().split("\t")):
        if i < len(headers):
            if headers[i] == "id_sense_wiki":
                test_sense_ids.append(el)

for sense_id in eval_sense_ids:
    if sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

for sense_id in train_sense_ids:
    if sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)

for sense_id in test_sense_ids:
    if sense_id in eval_sense_ids:
        supp_sense_ids.append(sense_id)
    if sense_id in train_sense_ids:
        supp_sense_ids.append(sense_id)

eval_data.close()
train_data.close()
test_data.close()

print(len(supp_sense_ids))

print(len(train_sense_ids))
print(len(set(train_sense_ids)))

print(len(eval_sense_ids))
print(len(set(eval_sense_ids)))

print(len(test_sense_ids))
print(len(set(test_sense_ids)))
"""
"""
with open("train.pkl", 'rb') as file:
    train_examples = pickle.load(file)
with open("dev.pkl", 'rb') as file:
    dev_examples = pickle.load(file)
with open("test_2.pkl", 'rb') as file:
    test_examples = pickle.load(file)

train_senses = []
dev_senses = []
test_senses = []

for tr in train_examples:
    train_senses.append(tr["definition"])
for dev in dev_examples:
    dev_senses.append(dev["definition"])
for test in test_examples:
    test_senses.append(test["definition"])

print(len(train_senses))
print(len(set(train_senses)))
print(len(dev_senses))
print(len(set(dev_senses)))
print(len(test_senses))
print(len(set(test_senses)))

import collections
print([item for item, count in collections.Counter(train_senses).items() if count > 1])
"""
"""
import pandas as pd
import pickle

def create_dataframe_from_pickle(pickle_file):
    try:
        # Read the list of dictionaries from the pickle file
        with open(pickle_file, 'rb') as file:
            data = pickle.load(file)

        # Convert the list of dictionaries into a DataFrame
        df = pd.DataFrame(data)

        return df
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# Usage example:
pickle_file = 'dev.pkl'  # Replace with your pickle file path
data_frame = create_dataframe_from_pickle(pickle_file)

if data_frame is not None:
    # Now you can work with the data_frame
    print(data_frame.head())  # Print the first few rows of the DataFrame


print(data_frame["supersense"].value_counts())
print(data_frame["supersense"].value_counts().sum())
"""
"""
import pandas as pd
import pickle

# Define the function to create a DataFrame from a pickle file
def create_dataframe_from_pickle(pickle_file):
    try:
        with open(pickle_file, 'rb') as file:
            data = pickle.load(file)
        df = pd.DataFrame(data)
        return df
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# List of file names for your data
data_files = ['train.pkl', 'dev.pkl', 'test.pkl', 'test_2.pkl']

# Create an empty DataFrame to store the combined distribution
combined_distribution = pd.DataFrame(columns=["File", "Supersense", "Count"])

# Iterate through the data files and calculate the supersense distribution for each
for file_name in data_files:
    # Load the data into a DataFrame
    data_frame = create_dataframe_from_pickle(file_name)

    if data_frame is not None:
        # Calculate the supersense distribution for the current file
        distribution = data_frame["supersense"].value_counts().reset_index()
        print(data_frame["supersense"].value_counts().sum())
        distribution.columns = ["Supersense", "Count"]
        # Remove ".pkl" from the file name in the "File" column
        distribution["File"] = file_name.replace('.pkl', '')
        combined_distribution = pd.concat([combined_distribution, distribution], ignore_index=True)

# Save the combined distribution to an Excel file
excel_filename = 'combined_supersense_distribution.xlsx'
combined_distribution.to_excel(excel_filename, sheet_name='Combined Distribution', index=False)
"""

